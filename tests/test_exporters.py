from pathlib import Path

from openpyxl import load_workbook

from jobpicky.exporters import build_feishu_record, export_jobs_to_excel


def test_build_feishu_record_uses_alljobs_field_names():
    row = {
        "source": "WonderCV",
        "company": "示例公司",
        "title": "FPGA工程师",
        "priority": "push",
        "is_relevant": 1,
        "match_score": 90,
        "detail_url": "https://example.com/job",
        "official_url": "https://careers.example.com/job",
    }

    record = build_feishu_record(row)
    assert record["fields"]["官方链接"]["link"] == "https://careers.example.com/job"

    assert record["fields"]["公司"] == "示例公司"
    assert "是否关注" not in record["fields"]
    assert "匹配分数" not in record["fields"]
    assert record["fields"]["原始链接"]["link"] == "https://example.com/job"


def test_export_jobs_to_excel_writes_headers_and_rows(tmp_path: Path):
    output = tmp_path / "jobs.xlsx"

    export_jobs_to_excel(
        [
            {
                "source": "WonderCV",
                "company": "示例公司",
                "title": "FPGA工程师",
                "priority": "push",
                "match_score": 90,
            }
        ],
        output,
    )

    wb = load_workbook(output)
    ws = wb.active
    assert "官方链接" in [cell.value for cell in ws[1]]
    assert ws["A1"].value == "岗位ID"
    assert ws["B2"].value == "示例公司"


def test_normalize_uses_user_status_field():
    from jobpicky.exporters import _normalize_all_job_row
    row = {
        "id": 12,
        "company": "TargetCo",
        "title": "研发",
        "user_status": None,
        "note": None
    }
    res = _normalize_all_job_row(row)
    assert res["user_status"] == "未看"
    assert res["note"] == ""


